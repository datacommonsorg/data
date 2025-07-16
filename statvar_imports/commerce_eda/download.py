import requests
import os
from urllib.parse import urlparse, urljoin
import openpyxl
import csv
import sys
from absl import logging
import shutil
from bs4 import BeautifulSoup


# Common HTTP headers to mimic a browser
COMMON_BROWSER_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Encoding': 'gzip, deflate, br',
    'Accept-Language': 'en-US,en;q=0.9',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'none',
    'Sec-Fetch-User': '?1',
}


def get_download_url_from_html(landing_page_url: str, link_criteria: dict) -> str | None:
    """Visits a landing page, parses its HTML, and extracts the full download URL."""
    logging.info(f"Attempting to find download URL on: {landing_page_url}")
    
    headers = COMMON_BROWSER_HEADERS.copy()
    headers['Referer'] = landing_page_url

    try:
        response = requests.get(landing_page_url, headers=headers, timeout=30)
        response.raise_for_status() # Raise HTTPError for bad responses
        
        soup = BeautifulSoup(response.text, 'html.parser')

        found_href = None
        
        bs4_attrs = {}
        if 'class' in link_criteria:
            bs4_attrs['class'] = link_criteria['class']
        if 'title' in link_criteria:
            bs4_attrs['title'] = link_criteria['title']

        links = soup.find_all('a', attrs=bs4_attrs)

        for link in links:
            if 'text_contains' in link_criteria:
                if link_criteria['text_contains'] not in link.get_text(strip=True):
                    continue
            
            if link.get('href'):
                found_href = link.get('href')
                break
        
        if found_href:
            full_download_url = urljoin(landing_page_url, found_href)
            logging.info(f"Found direct download URL: {full_download_url}")
            return full_download_url
        else:
            logging.error(f"Could not find download link on {landing_page_url} with criteria: {link_criteria}")
            return None

    except requests.exceptions.RequestException as e:
        logging.error(f"Error accessing landing page {landing_page_url}: {e}")
        return None
    except Exception as e:
        logging.error(f"An unexpected error occurred while parsing {landing_page_url}: {e}")
        return None


def download_file(url: str, output_folder: str) -> str | None:
    """Downloads a file from a given URL to a specified output folder."""
    try:
        os.makedirs(output_folder, exist_ok=True)
        parsed_url = urlparse(url)
        filename = os.path.basename(parsed_url.path)
        if not filename:
            logging.error(f"Could not determine filename from URL: {url}")
            return None

        save_path = os.path.join(output_folder, filename)
        logging.info(f"Attempting to download: {url} to {save_path}")

        headers = COMMON_BROWSER_HEADERS.copy()
        headers['Referer'] = url 

        response = requests.get(url, headers=headers, stream=True, timeout=30)
        response.raise_for_status() # Raise HTTPError for bad responses

        with open(save_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)

        logging.info(f"Successfully downloaded: {filename}")
        return save_path

    except requests.exceptions.RequestException as e:
        logging.error(f"Error downloading {url}: {e}")
    except IOError as e:
        logging.error(f"Error writing file to disk at {save_path}: {e}")
    except Exception as e:
        logging.error(f"An unexpected error occurred during download: {e}")
    return None

def extract_sheets_to_csv(excel_file_path: str, sheets_info: list[dict], output_folder: str) -> bool:
    """Extracts specified sheets from an XLSX file and saves them as separate CSV files with specified names."""
    if not os.path.exists(excel_file_path):
        logging.error(f"Excel file not found at {excel_file_path}")
        return False

    try:
        logging.info(f"\nLoading workbook: {excel_file_path}")
        workbook = openpyxl.load_workbook(excel_file_path, data_only=True) 

        all_sheet_names = workbook.sheetnames
        logging.info(f"Available sheets in '{os.path.basename(excel_file_path)}': {all_sheet_names}")

        os.makedirs(output_folder, exist_ok=True)
        logging.info(f"CSV files will be saved to: {os.path.abspath(output_folder)}") 

        all_sheets_extracted_successfully = True
        for sheet_spec in sheets_info:
            sheet_name = sheet_spec["name"]
            output_csv_name = sheet_spec["output_csv_name"]

            if sheet_name in all_sheet_names:
                logging.info(f"Extracting sheet: '{sheet_name}' as '{output_csv_name}'")
                sheet = workbook[sheet_name]
                
                csv_save_path = os.path.join(output_folder, output_csv_name)

                has_actual_content_written = False
                try:
                    with open(csv_save_path, 'w', newline='', encoding='utf-8') as csv_file:
                        csv_writer = csv.writer(csv_file)
                        
                        logging.debug(f"Starting write for CSV: {csv_save_path}")
                        row_counter = 0
                        for row_data in sheet.iter_rows(values_only=True):
                            row_counter += 1
                            if row_counter <= 5: 
                                logging.debug(f"Sheet '{sheet_name}' - Row {row_counter}: {row_data}")
                            
                            # Write row if it's the header or if it contains any non-empty cells
                            if row_counter == 1:
                                csv_writer.writerow(row_data)
                                if any(cell is not None for cell in row_data):
                                     has_actual_content_written = True
                            elif any(cell is not None and str(cell).strip() != '' for cell in row_data):
                                csv_writer.writerow(row_data)
                                has_actual_content_written = True
                        
                        logging.debug(f"Finished iterating rows for '{sheet_name}'. Total rows read: {row_counter}")

                    if not has_actual_content_written:
                        logging.warning(f"Sheet '{sheet_name}' was empty or contained only empty cells. No meaningful data written to CSV: {csv_save_path}")
                    
                    logging.info(f"Saved sheet '{sheet_name}' to: {csv_save_path}")

                except IOError as e:
                    logging.error(f"IOError writing CSV for sheet '{sheet_name}' to {csv_save_path}: {e}")
                    all_sheets_extracted_successfully = False
                except Exception as e:
                    logging.error(f"Unexpected error during CSV writing for sheet '{sheet_name}': {e}")
                    all_sheets_extracted_successfully = False

            else:
                logging.warning(f"Sheet '{sheet_name}' not found in '{os.path.basename(excel_file_path)}'. Skipping.")
                all_sheets_extracted_successfully = False
        
        return all_sheets_extracted_successfully

    except openpyxl.utils.exceptions.InvalidFileException:
        logging.error(f"Error: {excel_file_path} is not a valid Excel file or is corrupted.")
    except Exception as e:
        logging.error(f"An unexpected error occurred during sheet extraction from {excel_file_path}: {e}")
    return False


if __name__ == "__main__":
    # Determine the base directory of the script
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

    files_to_process = [
        {
            "landing_page_url": "https://www.eda.gov/performance/data-disclaimer",
            "link_criteria": {"text_contains": "I understand (XLSX)", "class": ["button"]},
            "sheets_info": [
                {"name": "Investments by State, Program", "output_csv_name": "Investment.csv"},
                {"name": "Estimated Outcomes by State", "output_csv_name": "EstimatedOutcome.csv"}
            ]
        },
        {
            "landing_page_url": "https://www.eda.gov/performance/tools",
            "link_criteria": {"text_contains": "Persistent Poverty Counties", "title": "Persistent Poverty Counties"},
            "sheets_info": [
                {"name": "Underlying_Data", "output_csv_name": "Poverty.csv"}
            ]
        }
    ]

    # Define output folders relative to the script's directory
    download_base_folder = os.path.join(SCRIPT_DIR, "downloaded_eda_data")
    extracted_csv_base_folder = os.path.join(SCRIPT_DIR, "input_file")
    processed_files_folder = os.path.join(download_base_folder, "processed")

    logging.info("--- Starting download and extraction process ---")

    overall_success = True

    for file_info in files_to_process:
        landing_url = file_info["landing_page_url"]
        link_crit = file_info["link_criteria"]
        sheets_to_extract_info = file_info["sheets_info"] # Use the new key

        logging.info(f"\nProcessing landing URL: {landing_url}")
        
        download_url = get_download_url_from_html(landing_url, link_crit)

        if download_url:
            downloaded_file_path = download_file(download_url, download_base_folder)

            if downloaded_file_path:
                extraction_success = extract_sheets_to_csv(downloaded_file_path, 
                                                            sheets_to_extract_info, 
                                                            extracted_csv_base_folder)
                
                if not extraction_success:
                    logging.error(f"Failed to extract all specified sheets from {os.path.basename(downloaded_file_path)}. Check logs above.")
                    overall_success = False
                else:
                    logging.info(f"Successfully extracted all specified sheets from {os.path.basename(downloaded_file_path)}.")
                    
                    try:
                        os.makedirs(processed_files_folder, exist_ok=True)
                        destination_path = os.path.join(processed_files_folder, os.path.basename(downloaded_file_path))
                        shutil.move(downloaded_file_path, destination_path)
                        logging.info(f"Moved original downloaded file to processed folder: {destination_path}")
                    except Exception as e:
                        logging.warning(f"Could not move original downloaded file {downloaded_file_path} to processed folder: {e}")
            else:
                logging.error(f"Download failed for {download_url}. Skipping sheet extraction.")
                overall_success = False
        else:
            logging.error(f"Could not determine direct download URL from landing page {landing_url}. Skipping this file.")
            overall_success = False
        
        logging.info("-" * 50)

    if overall_success:
        logging.info("\nAll specified files and sheets processed successfully! 🎉")
    else:
        logging.fatal("\nProcess completed with errors for some files/sheets. Please review the logs.")
        sys.exit(1)
